import csv
import json
import math
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


DATE_FIELDS = {"Trade Date", "Expiry Date"}
TEXT_FIELDS = {"Stock Symbol", "Call Option Recommendation", "Put Option Recommendation", "Confidence Level"}
NEAR_MONEY_STRIKES_PER_SIDE = 10
NEAR_MONEY_SHEET_NAME = "10 ITM OTM"
RECOMMENDATION_FIELDS = [
    "Call Option Recommendation",
    "Put Option Recommendation",
    "Confidence Level",
]


def parse_date(value):
    return datetime.strptime(value, "%Y-%m-%d").date() if value else None


def parse_float(value):
    return float(value) if value not in ("", None) else None


def parse_int(value):
    return int(float(value)) if value not in ("", None) else None


def parse_value(header, value):
    if value in ("", None):
        return None
    if header in DATE_FIELDS:
        return parse_date(value)
    if header in TEXT_FIELDS:
        return value
    return parse_float(value)


def confidence_label(score):
    if score >= 75:
        return "High"
    if score >= 55:
        return "Medium"
    if score >= 40:
        return "Low"
    return "N/A"


def recommendation_score(row, side, future_price, trend_pct, max_values):
    prefix = "Call" if side == "call" else "Put"
    oi = parse_float(row.get(f"{prefix} OI")) or 0
    volume = parse_float(row.get(f"{prefix} Volume")) or 0
    change_oi = max(parse_float(row.get(f"{prefix} Change in OI")) or 0, 0)
    price = parse_float(row.get(f"{prefix} Price")) or 0
    delta = abs(parse_float(row.get(f"{prefix} Delta")) or 0)
    strike = parse_float(row.get("Strike Price"))
    if price <= 0 or oi <= 0:
        return 0

    oi_score = math.log1p(oi) / math.log1p(max_values["oi"]) if max_values["oi"] > 0 else 0
    volume_score = math.log1p(volume) / math.log1p(max_values["volume"]) if max_values["volume"] > 0 else 0
    change_score = math.log1p(change_oi) / math.log1p(max_values["change_oi"]) if max_values["change_oi"] > 0 else 0
    delta_fit = max(0, 1 - abs(delta - 0.45) / 0.45)
    moneyness_fit = 0
    if future_price and strike:
        moneyness_fit = 1 / (1 + abs(strike - future_price) / future_price * 10)
    if trend_pct is None:
        trend_fit = 0.5
    elif side == "call":
        trend_fit = min(max(trend_pct, 0) / 0.03, 1)
    else:
        trend_fit = min(max(-trend_pct, 0) / 0.03, 1)

    return round(
        100
        * (
            0.25 * oi_score
            + 0.20 * volume_score
            + 0.15 * change_score
            + 0.15 * delta_fit
            + 0.15 * moneyness_fit
            + 0.10 * trend_fit
        ),
        1,
    )


def styled_header(ws, labels):
    fill = PatternFill("solid", fgColor="174E63")
    font = Font(color="FFFFFF", bold=True)
    cells = []
    for label in labels:
        cell = WriteOnlyCell(ws, value=label)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")
        cells.append(cell)
    ws.append(cells)


def select_near_money_rows(input_csv):
    groups = {}
    symbol_date_prices = {}
    with input_csv.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for index, row in enumerate(reader):
            future_price = parse_float(row.get("Future Price"))
            strike_price = parse_float(row.get("Strike Price"))
            if future_price is None or strike_price is None:
                continue
            key = (row.get("Trade Date"), row.get("Stock Symbol"), row.get("Expiry Date"))
            groups.setdefault(key, []).append((index, row, future_price, strike_price))
            symbol_date_prices.setdefault((row.get("Stock Symbol"), row.get("Trade Date")), future_price)

    previous_prices = {}
    by_symbol = {}
    for (symbol, trade_date), future_price in symbol_date_prices.items():
        by_symbol.setdefault(symbol, []).append((trade_date, future_price))
    for symbol, prices in by_symbol.items():
        prices.sort()
        previous = None
        for trade_date, future_price in prices:
            previous_prices[(symbol, trade_date)] = previous
            previous = future_price

    selected = []
    for key in sorted(groups):
        group = groups[key]
        below = sorted(
            (item for item in group if item[3] < item[2]),
            key=lambda item: (item[2] - item[3], item[0]),
        )[:NEAR_MONEY_STRIKES_PER_SIDE]
        above = sorted(
            (item for item in group if item[3] > item[2]),
            key=lambda item: (item[3] - item[2], item[0]),
        )[:NEAR_MONEY_STRIKES_PER_SIDE]
        group_selected = sorted(below + above, key=lambda item: (item[3], item[0]))
        highlight_indexes = {item[0] for item in below[:2] + above[:2]}

        trade_date, symbol, _expiry = key
        previous_price = previous_prices.get((symbol, trade_date))
        trend_pct = None
        if previous_price:
            trend_pct = (group_selected[0][2] - previous_price) / previous_price

        max_values = {
            "oi": max(
                [parse_float(item[1].get("Call OI")) or 0 for item in group_selected]
                + [parse_float(item[1].get("Put OI")) or 0 for item in group_selected]
                + [0]
            ),
            "volume": max(
                [parse_float(item[1].get("Call Volume")) or 0 for item in group_selected]
                + [parse_float(item[1].get("Put Volume")) or 0 for item in group_selected]
                + [0]
            ),
            "change_oi": max(
                [max(parse_float(item[1].get("Call Change in OI")) or 0, 0) for item in group_selected]
                + [max(parse_float(item[1].get("Put Change in OI")) or 0, 0) for item in group_selected]
                + [0]
            ),
        }

        scored = []
        for item in group_selected:
            row = dict(item[1])
            row["_highlight_nearest"] = item[0] in highlight_indexes
            call_score = recommendation_score(row, "call", item[2], trend_pct, max_values)
            put_score = recommendation_score(row, "put", item[2], trend_pct, max_values)
            scored.append((item, row, call_score, put_score))

        best_call = max(scored, key=lambda item: item[2], default=None)
        best_put = max(scored, key=lambda item: item[3], default=None)
        best_call_index = best_call[0][0] if best_call and best_call[2] >= 40 else None
        best_put_index = best_put[0][0] if best_put and best_put[3] >= 40 else None

        for item, row, call_score, put_score in scored:
            confidence_parts = []
            if item[0] == best_call_index:
                row["Call Option Recommendation"] = "Buy Call"
                confidence_parts.append(f"Call {confidence_label(call_score)}")
            else:
                row["Call Option Recommendation"] = "No Buy"
            if item[0] == best_put_index:
                row["Put Option Recommendation"] = "Buy Put"
                confidence_parts.append(f"Put {confidence_label(put_score)}")
            else:
                row["Put Option Recommendation"] = "No Buy"
            row["Confidence Level"] = "; ".join(confidence_parts) if confidence_parts else "N/A"
            selected.append(row)

    return selected


def main():
    if len(sys.argv) != 4:
        raise SystemExit("Usage: build_nse_fo_workbook.py INPUT_CSV METADATA_JSON OUTPUT_XLSX")

    input_csv = Path(sys.argv[1])
    metadata_path = Path(sys.argv[2])
    output_xlsx = Path(sys.argv[3])
    metadata = json.loads(metadata_path.read_text(encoding="utf-8"))
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    with input_csv.open(newline="", encoding="utf-8") as handle:
        headers = next(csv.reader(handle))
    near_money_headers = headers + RECOMMENDATION_FIELDS

    wb = Workbook(write_only=True)
    summary = wb.create_sheet("Summary")
    data = wb.create_sheet("Combined Data")

    title_fill = PatternFill("solid", fgColor="174E63")
    label_fill = PatternFill("solid", fgColor="F8FAFC")
    header_fill = PatternFill("solid", fgColor="E0F2FE")
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title = WriteOnlyCell(summary, value="NSE F&O Bhavcopy")
    title.fill = title_fill
    title.font = Font(color="FFFFFF", bold=True, size=16)
    summary.append([title])
    summary.append([])

    downloaded_dates = [item["date"] for item in metadata["downloaded"]]
    summary_rows = [
        ["Field", "Value"],
        ["Included period", f"{metadata['start_date']} to {metadata['end_date']}"],
        ["Trading-day files included", len(metadata["downloaded"])],
        ["Combined data rows", metadata["row_count"]],
        ["Latest included trade date", downloaded_dates[-1] if downloaded_dates else ""],
        ["Source archive folder", metadata["source_url"]],
        ["Source file pattern", "BhavCopy_NSE_FO_0_0_0_YYYYMMDD_F_0000.csv.zip"],
        ["Join logic", "STF futures joined to STO CE/PE rows by Trade Date + Stock Symbol + Expiry Date + Strike Price"],
        ["Price field", "Close price (ClsPric)"],
        ["OI field", "Open interest (OpnIntrst)"],
        ["Filtered rows", "Rows where both Call OI and Put OI are zero or blank are excluded."],
        [f"{NEAR_MONEY_SHEET_NAME} workbook", f"Created as a separate Excel file with the {NEAR_MONEY_STRIKES_PER_SIDE} nearest strikes below and {NEAR_MONEY_STRIKES_PER_SIDE} nearest strikes above the futures close for each Trade Date + Stock Symbol + Expiry Date. The nearest 2 strikes below and nearest 2 strikes above the futures close are highlighted in yellow."],
        ["Recommendation logic", "The separate ITM/OTM workbook marks one Buy Call and one Buy Put candidate per Trade Date + Stock Symbol + Expiry Date when score is at least 40. Scores use futures trend, option OI, volume, positive OI change, moneyness, and delta fit. This is a quantitative screen, not financial advice."],
        ["IV availability", "Strike-level Call IV and Put IV were not found in the local NSE bhavcopy, Contract_Delta, or FOVOLT source files. FOVOLT contains symbol-level volatility only."],
    ]
    for note_index, note in enumerate(metadata.get("source_notes", []), start=1):
        summary_rows.append([f"Source note {note_index}", note])
    for row_index, row in enumerate(summary_rows, start=1):
        out = []
        for col_index, value in enumerate(row):
            cell = WriteOnlyCell(summary, value=value)
            cell.border = border
            if row_index == 1:
                cell.fill = header_fill
                cell.font = Font(bold=True)
            elif col_index == 0:
                cell.fill = label_fill
                cell.font = Font(bold=True)
            if isinstance(value, str) and len(value) > 55:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            out.append(cell)
        summary.append(out)

    summary.append([])
    styled_header(summary, ["Included Trade Dates"])
    for included in downloaded_dates:
        summary.append([included])
    summary.append([])
    styled_header(summary, ["Skipped Dates", "Reason"])
    for skipped in metadata["skipped"]:
        summary.append([skipped["date"], skipped["reason"]])

    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 96
    summary.column_dimensions["C"].width = 4
    summary.column_dimensions["D"].width = 20

    data.freeze_panes = "A2"
    last_col = get_column_letter(len(headers))
    data.auto_filter.ref = f"A1:{last_col}{metadata['row_count'] + 1}"
    near_money_rows = select_near_money_rows(input_csv)
    for idx, header in enumerate(headers, start=1):
        letter = get_column_letter(idx)
        if header in DATE_FIELDS:
            width = 12
        elif header in TEXT_FIELDS:
            width = 16
        elif "Volatility" in header or "Change %" in header:
            width = 18
        else:
            width = 14
        data.column_dimensions[letter].width = width

    styled_header(data, headers)
    with input_csv.open(newline="", encoding="utf-8") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            data.append([parse_value(header, row.get(header)) for header in headers])

    wb.save(output_xlsx)

    near_money_xlsx = output_xlsx.with_name(f"{output_xlsx.stem}_10_itm_otm.xlsx")
    near_wb = Workbook(write_only=True)
    near_money = near_wb.create_sheet(NEAR_MONEY_SHEET_NAME)
    highlight_fill = PatternFill("solid", fgColor="FFF2CC")
    buy_call_fill = PatternFill("solid", fgColor="D9EAD3")
    buy_put_fill = PatternFill("solid", fgColor="F4CCCC")
    near_money.freeze_panes = "A2"
    near_money_last_col = get_column_letter(len(near_money_headers))
    near_money.auto_filter.ref = f"A1:{near_money_last_col}{len(near_money_rows) + 1}"
    for idx, header in enumerate(near_money_headers, start=1):
        letter = get_column_letter(idx)
        if header in DATE_FIELDS:
            width = 12
        elif header in RECOMMENDATION_FIELDS:
            width = 24
        elif header in TEXT_FIELDS:
            width = 16
        elif "Volatility" in header or "Change %" in header:
            width = 18
        else:
            width = 14
        near_money.column_dimensions[letter].width = width

    styled_header(near_money, near_money_headers)
    for row in near_money_rows:
        values = []
        row_fill = None
        if row.get("Call Option Recommendation") == "Buy Call":
            row_fill = buy_call_fill
        if row.get("Put Option Recommendation") == "Buy Put":
            row_fill = buy_put_fill
        for header in near_money_headers:
            cell = WriteOnlyCell(near_money, value=parse_value(header, row.get(header)))
            if row_fill:
                cell.fill = row_fill
            elif row.get("_highlight_nearest"):
                cell.fill = highlight_fill
            values.append(cell)
        near_money.append(values)
    near_wb.save(near_money_xlsx)

    check = load_workbook(output_xlsx, read_only=True, data_only=True)
    ws = check["Combined Data"]
    first_rows = list(ws.iter_rows(min_row=1, max_row=4, values_only=True))
    print(
        json.dumps(
            {
                "output": str(output_xlsx),
                "near_money_output": str(near_money_xlsx),
                "sheets": check.sheetnames,
                "data_rows_in_metadata": metadata["row_count"],
                "worksheet_max_row": metadata["row_count"] + 1,
                "worksheet_max_column": len(headers),
                "near_money_rows": len(near_money_rows),
                "near_money_columns": len(near_money_headers),
                "sample": first_rows,
            },
            default=str,
            indent=2,
        )
    )
    check.close()


if __name__ == "__main__":
    main()
